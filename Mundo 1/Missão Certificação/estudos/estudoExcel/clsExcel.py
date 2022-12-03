import openpyxl


def createInsertXLSX():

   book=openpyxl.Workbook()

   print(book.sheetnames)

   book.create_sheet('frutas')

   frutas_page=book['frutas']
   frutas_page.append(['fruta','quantidade','preco'])
   frutas_page.append(['bananas',5,'R$ 3.90'])
   frutas_page.append(['abacaxi',1,'R$ 6.50'])
   frutas_page.append(['maças',3,'R$ 4.00'])
   frutas_page.append(['caki',2,'R$ 8.10'])

   book.save('compras.xlsx')

def OpenReadXLSX():
    book=openpyxl.load_workbook('compras.xlsx') 
    frutas_page=book['frutas']
    
    """
    for rows in frutas_page.iter_rows(min_row=2,max_row=5):
        for cell in rows:
            #print(cell.value)
    """

    for rows in frutas_page.iter_rows(min_row=2,max_row=5):
        #print(rows[0].value,rows[1].value,rows[2].value)
        print(f'{rows[0].value},{rows[1].value},{rows[2].value}')   
             

def OpenUpDateXLSX():
    book=openpyxl.load_workbook('compras.xlsx') 
    frutas_page=book['frutas']
    
    for rows in frutas_page.iter_rows(min_row=2,max_row=5):
        for cell in rows:
            if cell.value=='caki':
                cell.value='Cachaça'


    book.save('compras.xlsx')

OpenUpDateXLSX()             