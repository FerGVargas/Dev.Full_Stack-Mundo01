import openpyxl
from pathlib import Path

 
class Dados():
    def __init__(self,pathDestino=''):

        # o formato esperado para pathlib tratar
        # o path e com a contrabarra mas o paramtro
        # não precisa
 
        if (pathDestino==''):
           self.pathDestino='planilhas/'
        else:    
           self.pathDestino=pathDestino + '/'
    
    # como é esperado o path das planilhas ou arquivos img 
    # por defaulr sem parametro será pathDestino='planilhas/'

    def getPathPlanilhas(self,nomeArquivo,sPath=''):
        """
          Tratamento necessário para trabalhar com subpastas
          Retorna o path com o nome do arquivo xlsx,img..

        """
        if (sPath!=''):
            data_folder = Path(sPath)
            new_path = data_folder / nomeArquivo
        else:
            data_folder = Path(self.pathDestino)
            new_path = data_folder / nomeArquivo  

        return new_path

    def openpyXL(self, nome_e_Extensao, nome_planilha):

        pathDestinoArq = self.getPathPlanilhas(nome_e_Extensao,self.pathDestino)

        try:
            book = openpyxl.load_workbook(pathDestinoArq)
        except:
            book = openpyxl.Workbook()
            book.create_sheet(nome_planilha)
        finally:
            return book

    def createInsertXLSX(self, nome_e_Extensao, nomePlanilha, dados):
        
        book = self.openpyXL(nome_e_Extensao, nomePlanilha)

        infoCells = book[nomePlanilha]

        infoCells.append(dados)
        
        #a função save é da bibioteca por isso precisa do tratamento do path
        pathDestinoArq  = self.getPathPlanilhas(nome_e_Extensao,self.pathDestino)

        book.save(pathDestinoArq)

    def OpenReadXLSX(self, nome_e_Extensao, nomePlanilha,lsCab=[]):
        #RETORNA UMA LISTA DE LISTAS DE REGISTROS
        #MATRIZ
        #CADA LINHA É UM REGISTRO OU LINHA DA PLANILHA
        book = self.openpyXL(nome_e_Extensao,nomePlanilha)
        infoCells = book[nomePlanilha]
        
        #DEBUG
        #print(' cols   ====>>',len(lsCab))
        
        nCols=len(lsCab)

        lsReturn = []
        try:
           #REDUZIR O NUMERO DE LINHAS DO GRID / para testar retorno usa max_row=1
           #min_row=1
           for rows in infoCells.iter_rows():
            
               ls=''
             
               for i in range(0,nCols):
                   if (len(ls)>0):
                      ls+= ','+rows[i].value
                   else:
                      ls= rows[i].value 

               if (rows[0].value!=None):
                  lsReturn.append(ls.upper())
               else: 
                  break

               #print('dados ========>>>>>',ls.split(','))
        except:
            pass 
        return lsReturn     

    def getList(self, nome_e_Extensao, nomePlanilha,col=0):

        # retorna uma lista lida da coluna 'A' da planilha
        book = self.openpyXL(nome_e_Extensao,nomePlanilha)
        infoCells = book[nomePlanilha]

        lsReturn = []

        for rows in infoCells.iter_rows(min_row=1):
            # print(rows[0].value,rows[1].value,rows[2].value)
            # print(f'{rows[0].value},{rows[1].value},{rows[2].value}')
            if (rows[col].value!=None):
              lsReturn.append(rows[col].value)
            else: 
              break
             
        return lsReturn

    def OpenFindDateXLSX(self, nome_e_Extensao, nomePlanilha, valor, nCols):
          
        book = self.openpyXL(nome_e_Extensao,nomePlanilha)
        infoCells = book[nomePlanilha]

        bReturn = False
        ls=[]

        for rows in infoCells.iter_rows(min_row=1):
            for i in range(0,nCols+1):
               print('rows[i].value-->>',rows[i].value) 
               if rows[i].value.upper()==valor.upper():
                    for i in range(0,nCols+1):
                        ls.append(rows[i].value)
        return ls               
                    
    
    def saveFileTemp(self,ls):
        print('saveFileTemp-->>>',ls) 
        arquivoTemp=open('newfileLista.txt','w')
        for i in range(0,len(ls)):
            if type(ls[i])==int:
                ls[i]=str(ls[i])
            arquivoTemp.write(ls[i]+'\n') 
        arquivoTemp.close()
        
    # How to read a file
    def readFileTemp(self):
        
        ls=[]
        sTemp=''
        try:
           arquivoTemp=open('newfileLista.txt','r')
           for line in arquivoTemp.readlines():
               sTemp=line[:-1]
               ls.append(sTemp)
        
           arquivoTemp.close()
        except:
            pass
        
        return(ls)

