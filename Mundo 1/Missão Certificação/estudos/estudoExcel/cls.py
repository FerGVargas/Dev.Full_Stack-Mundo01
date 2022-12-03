import openpyxl

class Dados1():
    def __init__(self):
        pass

    def openpyXL(self,nome_e_Extensao,nome_planilha):
        try:
           book=openpyxl.load_workbook(nome_e_Extensao)
        except:
           book=openpyxl.Workbook()
           book.create_sheet(nome_planilha)
        finally:
           return book

    def createInsertXLSX(self,nomeArquivo,nomePlanilha,dados):
 
         book=self.openpyXL(nomeArquivo,nomePlanilha)

         infoCells=book[nomePlanilha]
        
         infoCells.append(dados)
 
         book.save(nomeArquivo)

    def OpenReadXLSX(self,nomePlanilha,nomeArquivo):
        book=self.openpyXL(nomeArquivo,nomeArquivo) 
        infoCells=self.book[nomePlanilha]
        
 
        for rows in infoCells.iter_rows(min_row=2,max_row=5):
            #print(rows[0].value,rows[1].value,rows[2].value)
            print(f'{rows[0].value},{rows[1].value},{rows[2].value}')   



    def OpenFindDateXLSX(self,nomeArquivo,nomePlanilha,chave,valor):
        book=openpyxl.load_workbook(nomePlanilha) 
        infoCells=book[nomeArquivo]
        
        bReturn = False

        for rows in infoCells.iter_rows(min_row=2,max_row=5):
            for cell in rows:
                if cell.value==chave:
                    cell.value=valor
                    bReturn= True
                    
            return bReturn

        

    def OpenUpDateXLSX(self,nomePlanilha,nomeArquivo,chave,valor):
        book=openpyxl.load_workbook(nomePlanilha) 
        frutas_page=book[nomeArquivo]
        
        for rows in frutas_page.iter_rows(min_row=2,max_row=5):
            for cell in rows:
                if cell.value==chave:
                    cell.value=valor


        book.save(nomePlanilha)

              